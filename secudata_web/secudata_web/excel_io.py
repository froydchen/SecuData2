from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

from .models import FinalRecord, MeasurementMetadata, MeasurementRaw

HEADERS = [
    "createdAt",
    "measurementTimestamp",
    "rpe",
    "rins",
    "ipe",
    "u",
    "isOk",
    "id",
    "geraeteart",
    "hersteller",
    "raumEtage",
    "kunde",
    "typModell",
    "seriennummer",
    "pruefer",
    "zusatztext",
]


def export_records_to_bytes(records: list[FinalRecord]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messdaten"
    sheet.append(HEADERS)

    for record in records:
        sheet.append(
            [
                record.created_at,
                record.measurement.timestamp,
                record.measurement.rpe,
                record.measurement.rins,
                record.measurement.ipe,
                record.measurement.u,
                record.measurement.is_ok,
                record.metadata.id,
                record.metadata.geraeteart,
                record.metadata.hersteller,
                record.metadata.raum_etage,
                record.metadata.kunde,
                record.metadata.typ_modell,
                record.metadata.seriennummer,
                record.metadata.pruefer,
                record.metadata.zusatztext,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_records_to_path(records: list[FinalRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(export_records_to_bytes(records))


def import_records_from_bytes(payload: bytes) -> list[FinalRecord]:
    workbook = load_workbook(BytesIO(payload), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    if header[: len(HEADERS)] != HEADERS:
        raise ValueError("Excel-Kopf passt nicht zum erwarteten SECU-DAT-Format.")

    records: list[FinalRecord] = []
    for row in rows[1:]:
        if not row or all(value in (None, "") for value in row[: len(HEADERS)]):
            continue
        values = list(row) + [None] * max(0, len(HEADERS) - len(row))
        record = FinalRecord(
            id=0,
            created_at=_string(values[0]),
            measurement=MeasurementRaw(
                rpe=_float_or_none(values[2]),
                rins=_float_or_none(values[3]),
                ipe=_float_or_none(values[4]),
                u=_float_or_none(values[5]),
                timestamp=_string(values[1]),
                is_ok=_bool(values[6]),
            ),
            metadata=MeasurementMetadata(
                id=_string(values[7]),
                geraeteart=_string(values[8]),
                hersteller=_string(values[9]),
                raum_etage=_string(values[10]),
                kunde=_string(values[11]),
                typ_modell=_string(values[12]),
                seriennummer=_string(values[13]),
                pruefer=_string(values[14]),
                zusatztext=_string(values[15]),
            ),
        )
        if not record.created_at or not record.measurement.timestamp:
            raise ValueError("Excel enthält Datensatz ohne Zeitstempel.")
        records.append(record)
    return records


def _string(value) -> str:
    return "" if value is None else str(value).strip()


def _float_or_none(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return float(text.replace(",", "."))


def _bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "wahr", "yes", "ja"}
